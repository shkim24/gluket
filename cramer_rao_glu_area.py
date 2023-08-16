import openpyxl

def process_excel_file(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Choose the sheets you want to work with (replace 'Sheet1' and 'OutputSheet' with actual sheet names)
    input_sheet = workbook['cramer_rao']
    output_sheet = workbook['glucr']
    
    # Get the maximum row and column counts in the input sheet
    max_row = input_sheet.max_row
    max_col = input_sheet.max_column
    
    # Iterate through each cell in the input sheet
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            input_cell = input_sheet.cell(row=row, column=col)
            if isinstance(input_cell.value, (int, float)) and input_cell.value > 20:
                # Delete the value and update the same cell location in the output sheet
                output_cell = output_sheet.cell(row=row, column=col)
                output_cell.value = None
    
    # Save changes
    workbook.save(file_path)
    print("Excel file processed successfully.")



excel_file_path = 'combined_sheet.xlsx'
process_excel_file(excel_file_path)
